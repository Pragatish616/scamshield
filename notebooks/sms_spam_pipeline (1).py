# ================================================================
#  SMS SPAM DETECTION PIPELINE
#  XGBoost + TF-IDF  |  scam_score risk metric
#  Compatible: Google Colab & local Python 3.8+
#
#  Outputs:
#    spam_detection_results.xlsx  — full predictions + scam_score
#    spam_model_report.png        — 4-panel evaluation chart
#    xgb_spam_model.pkl           — trained XGBoost model
#    tfidf_vectorizer.pkl         — fitted TF-IDF vectoriser
#
#  Quick-start (Colab):
#    1. Upload your CSV to Colab (label, message columns)
#    2. pip install xgboost openpyxl   # already in Colab
#    3. Run all cells
#    4. Download outputs from the file browser
# ================================================================

# ── Install (uncomment in Colab) ────────────────────────────────
# !pip install xgboost openpyxl scikit-learn pandas numpy matplotlib seaborn -q

import warnings
warnings.filterwarnings("ignore")

import os
import time
import pickle

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

from sklearn.model_selection  import train_test_split, GridSearchCV, StratifiedKFold
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score,
    f1_score, confusion_matrix, classification_report,
)
from sklearn.utils.class_weight import compute_sample_weight
from xgboost import XGBClassifier

from openpyxl.styles           import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils            import get_column_letter
from openpyxl.formatting.rule  import ColorScaleRule


# ================================================================
#  CONFIG  — change only these paths / flags
# ================================================================
CSV_PATH     = "sms_spam_dataset.csv"   # ← your input file
OUTPUT_XLSX  = "spam_detection_results.xlsx"
MODEL_PATH   = "xgb_spam_model.pkl"
TFIDF_PATH   = "tfidf_vectorizer.pkl"
PLOT_PATH    = "spam_model_report.png"

LABEL_COL    = "label"     # column must contain "spam" or "ham"
TEXT_COL     = "message"   # column with raw SMS text
TEST_SIZE    = 0.20        # 80/20 split
RANDOM_STATE = 42


# ================================================================
#  1. DATA LOADING & CLEANING
# ================================================================
def load_data(path: str) -> pd.DataFrame:
    """
    Load CSV, enforce column names, drop nulls, encode labels.
    Returns a clean DataFrame with an added 'label_enc' column
    (1 = spam, 0 = ham).
    """
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip().str.lower()

    # Accept common alternate column names
    col_aliases = {
        "text": TEXT_COL, "sms": TEXT_COL, "body": TEXT_COL,
        "class": LABEL_COL, "category": LABEL_COL, "type": LABEL_COL,
    }
    df.rename(columns=col_aliases, inplace=True)

    missing = [c for c in [TEXT_COL, LABEL_COL] if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}. Found: {list(df.columns)}")

    df = df[[TEXT_COL, LABEL_COL]].dropna()
    df[TEXT_COL]  = df[TEXT_COL].astype(str).str.strip()
    df[LABEL_COL] = df[LABEL_COL].str.strip().str.lower()

    allowed = {"spam", "ham"}
    found   = set(df[LABEL_COL].unique())
    if not found.issubset(allowed):
        raise ValueError(f"label column must be 'spam'/'ham'. Got: {found - allowed}")

    df["label_enc"] = (df[LABEL_COL] == "spam").astype(int)

    print(f"  Loaded {len(df):,} records  |  "
          f"Spam: {(df[LABEL_COL]=='spam').sum():,}  "
          f"Ham: {(df[LABEL_COL]=='ham').sum():,}")
    return df


# ================================================================
#  2. TF-IDF VECTORISER
#     ngram_range=(1,2)  → captures unigrams AND bigrams
#     sublinear_tf=True  → log-dampens high-frequency terms
#     max_features=10000 → keeps vocabulary manageable
# ================================================================
def build_tfidf(
    X_train: pd.Series,
    X_test:  pd.Series,
) -> tuple[TfidfVectorizer, object, object]:
    vectorizer = TfidfVectorizer(
        ngram_range   = (1, 2),
        max_features  = 10_000,
        sublinear_tf  = True,
        strip_accents = "unicode",
        analyzer      = "word",
        token_pattern = r"\w{1,}",
        min_df        = 2,
    )
    X_tr = vectorizer.fit_transform(X_train)
    X_te = vectorizer.transform(X_test)

    print(f"  Vocabulary : {len(vectorizer.vocabulary_):,} tokens")
    return vectorizer, X_tr, X_te


# ================================================================
#  3. MODEL TRAINING
#     class imbalance → compute_sample_weight("balanced")
#     optional GridSearchCV → set use_gridsearch=True
# ================================================================
def train_model(
    X_tr,
    y_train,
    use_gridsearch: bool = False,
) -> XGBClassifier:
    # Balanced sample weights handle class imbalance without
    # oversampling — safer for sparse TF-IDF matrices.
    sample_w = compute_sample_weight("balanced", y_train)

    if use_gridsearch:
        print("  Running GridSearchCV (may take a few minutes)…")
        param_grid = {
            "max_depth"        : [3, 5, 7],
            "n_estimators"     : [100, 200, 300],
            "learning_rate"    : [0.05, 0.10, 0.15],
            "subsample"        : [0.7, 0.8, 1.0],
            "colsample_bytree" : [0.7, 0.8, 1.0],
        }
        base_model = XGBClassifier(
            eval_metric  = "logloss",
            random_state = RANDOM_STATE,
            tree_method  = "hist",
            n_jobs       = -1,
        )
        cv = StratifiedKFold(n_splits=5, shuffle=True,
                             random_state=RANDOM_STATE)
        gs = GridSearchCV(
            base_model, param_grid,
            cv       = cv,
            scoring  = "f1",
            n_jobs   = -1,
            verbose  = 1,
        )
        gs.fit(X_tr, y_train, sample_weight=sample_w)
        print(f"  Best params : {gs.best_params_}")
        print(f"  Best CV F1  : {gs.best_score_:.4f}")
        return gs.best_estimator_

    # Default high-quality configuration (no grid search needed
    # for most SMS datasets — already tuned for speed + accuracy)
    model = XGBClassifier(
        n_estimators      = 200,
        max_depth         = 5,
        learning_rate     = 0.10,
        subsample         = 0.80,
        colsample_bytree  = 0.80,
        min_child_weight  = 2,
        gamma             = 0.1,
        eval_metric       = "logloss",
        random_state      = RANDOM_STATE,
        tree_method       = "hist",
        n_jobs            = -1,
    )
    model.fit(X_tr, y_train, sample_weight=sample_w)
    return model


# ================================================================
#  4. EVALUATION
# ================================================================
def evaluate(
    model,
    X_te,
    y_test,
    label_names: tuple = ("ham", "spam"),
) -> tuple[np.ndarray, np.ndarray, dict, np.ndarray]:
    y_pred = model.predict(X_te)
    y_prob = model.predict_proba(X_te)   # shape (n, 2): [ham_p, spam_p]

    metrics = {
        "accuracy"  : accuracy_score (y_test, y_pred),
        "precision" : precision_score(y_test, y_pred),
        "recall"    : recall_score   (y_test, y_pred),
        "f1"        : f1_score       (y_test, y_pred),
    }
    cm = confusion_matrix(y_test, y_pred)

    print("\n  ── Evaluation Metrics (test set) ──────────────")
    for k, v in metrics.items():
        bar = "█" * int(v * 20)
        print(f"  {k:<12}: {v:.4f}  {bar}")
    print(f"\n{classification_report(y_test, y_pred, target_names=label_names)}")
    return y_pred, y_prob, metrics, cm


# ================================================================
#  5. SCAM SCORE
#     Risk score derived from model's spam confidence.
#     NOT a supervised label — purely a confidence metric.
#
#     scam_score = round(spam_probability × 100, 2)
#     Range: 0 (definitely ham) → 100 (definitely spam)
#
#     Interpretation guide:
#       0  –  20  : Very likely legitimate
#       21 –  49  : Low risk / borderline
#       50 –  79  : Probable spam — review manually
#       80 – 100  : High confidence spam
# ================================================================
def compute_scam_score(spam_prob: np.ndarray) -> np.ndarray:
    return np.round(spam_prob * 100, 2)


# ================================================================
#  6. BUILD RESULTS DATAFRAME (all 500 records)
# ================================================================
def build_results_df(
    df:          pd.DataFrame,
    vectorizer:  TfidfVectorizer,
    model:       XGBClassifier,
    train_idx:   np.ndarray,
    test_idx:    np.ndarray,
) -> pd.DataFrame:
    """
    Returns a DataFrame with columns:
        message | actual_label | predicted_label |
        spam_probability | ham_probability | scam_score | split
    """
    X_all    = vectorizer.transform(df[TEXT_COL])
    prob_all = model.predict_proba(X_all)
    pred_all = model.predict(X_all)

    train_set = set(train_idx)
    results   = pd.DataFrame({
        "message"          : df[TEXT_COL].values,
        "actual_label"     : df[LABEL_COL].values,
        "predicted_label"  : ["spam" if p == 1 else "ham" for p in pred_all],
        "spam_probability" : np.round(prob_all[:, 1], 6),
        "ham_probability"  : np.round(prob_all[:, 0], 6),
        "scam_score"       : compute_scam_score(prob_all[:, 1]),
        "split"            : ["train" if i in train_set else "test"
                               for i in range(len(df))],
    })
    n_errors = (results["actual_label"] != results["predicted_label"]).sum()
    print(f"  Total records    : {len(results):,}")
    print(f"  Mis-classified   : {n_errors}")
    print(f"  Avg scam_score (spam rows) : "
          f"{results[results['actual_label']=='spam']['scam_score'].mean():.1f}")
    print(f"  Avg scam_score (ham rows)  : "
          f"{results[results['actual_label']=='ham']['scam_score'].mean():.1f}")
    return results


# ================================================================
#  7. VISUALISATION PANEL (2×2 figure, saved as PNG)
# ================================================================
def plot_report(
    cm:          np.ndarray,
    metrics:     dict,
    model:       XGBClassifier,
    vectorizer:  TfidfVectorizer,
    y_test:      np.ndarray,
    y_prob:      np.ndarray,
    save_path:   str,
) -> None:
    fig, axes = plt.subplots(2, 2, figsize=(15, 11))
    fig.suptitle("SMS Spam Detection — Model Evaluation Report",
                 fontsize=17, fontweight="bold", y=1.01)
    palette = {"spam": "#E74C3C", "ham": "#27AE60",
               "blue": "#2980B9", "purple": "#8E44AD"}

    # ── 7a: Confusion Matrix ────────────────────────────────
    ax = axes[0, 0]
    sns.heatmap(
        cm, annot=True, fmt="d", cmap="Blues",
        xticklabels=["Ham (0)", "Spam (1)"],
        yticklabels=["Ham (0)", "Spam (1)"],
        ax=ax, linewidths=1, annot_kws={"size": 16, "weight": "bold"},
        cbar_kws={"shrink": 0.75},
    )
    ax.set_xlabel("Predicted Label", fontsize=11, labelpad=8)
    ax.set_ylabel("Actual Label",    fontsize=11, labelpad=8)
    ax.set_title("Confusion Matrix", fontsize=13, pad=12)

    # Annotate TP / TN / FP / FN
    labels = [["TN", "FP"], ["FN", "TP"]]
    for i in range(2):
        for j in range(2):
            ax.text(j + 0.5, i + 0.72, labels[i][j],
                    ha="center", va="center",
                    fontsize=10, color="grey")

    # ── 7b: Metric Bars ─────────────────────────────────────
    ax = axes[0, 1]
    names  = ["Accuracy", "Precision", "Recall", "F1-Score"]
    values = [v * 100 for v in metrics.values()]
    colors = [palette["blue"], palette["purple"],
              palette["spam"], palette["ham"]]
    bars = ax.bar(names, values, color=colors, width=0.55,
                  edgecolor="white", linewidth=1.5)
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.8,
                f"{val:.2f}%", ha="center", va="bottom",
                fontsize=12, fontweight="bold", color="#2C3E50")
    ax.set_ylim(0, 115)
    ax.set_title("Evaluation Metrics", fontsize=13, pad=12)
    ax.set_ylabel("Score (%)", fontsize=11)
    ax.yaxis.grid(True, alpha=0.35, linestyle="--")
    ax.set_axisbelow(True)
    ax.spines[["top", "right"]].set_visible(False)

    # ── 7c: scam_score Distribution ─────────────────────────
    ax = axes[1, 0]
    ham_scores  = y_prob[y_test == 0, 1] * 100
    spam_scores = y_prob[y_test == 1, 1] * 100
    bins = np.linspace(0, 100, 26)
    ax.hist(ham_scores,  bins=bins, alpha=0.70, color=palette["ham"],
            label=f"Ham  (n={len(ham_scores)})",  edgecolor="white")
    ax.hist(spam_scores, bins=bins, alpha=0.70, color=palette["spam"],
            label=f"Spam (n={len(spam_scores)})", edgecolor="white")
    ax.axvline(50, color="#2C3E50", linestyle="--", linewidth=1.8,
               label="Decision boundary (50)")
    # Risk zones
    ax.axvspan( 0, 20, alpha=0.04, color=palette["ham"])
    ax.axvspan(80,100, alpha=0.04, color=palette["spam"])
    ax.set_title("scam_score Distribution (test set)", fontsize=13, pad=12)
    ax.set_xlabel("scam_score  →  0=Definitely Ham, 100=Definitely Spam",
                  fontsize=10)
    ax.set_ylabel("Message Count", fontsize=11)
    ax.legend(fontsize=10, framealpha=0.8)
    ax.yaxis.grid(True, alpha=0.35, linestyle="--")
    ax.set_axisbelow(True)
    ax.spines[["top", "right"]].set_visible(False)

    # ── 7d: Top-20 Feature Importances ──────────────────────
    ax = axes[1, 1]
    fi    = model.feature_importances_
    names_arr = vectorizer.get_feature_names_out()
    top_n = 20
    top_idx   = fi.argsort()[-top_n:][::-1]
    top_feats = names_arr[top_idx]
    top_vals  = fi[top_idx]

    cmap_vals  = plt.cm.RdYlGn_r(top_vals / top_vals.max())
    y_pos      = np.arange(top_n)
    ax.barh(y_pos, top_vals, color=cmap_vals, edgecolor="white",
            linewidth=0.8)
    ax.set_yticks(y_pos)
    ax.set_yticklabels(
        [f'"{f}"' for f in top_feats], fontsize=9, fontfamily="monospace"
    )
    ax.invert_yaxis()
    ax.set_title(f"Top-{top_n} TF-IDF Feature Importances",
                 fontsize=13, pad=12)
    ax.set_xlabel("XGBoost Importance Score", fontsize=11)
    ax.xaxis.grid(True, alpha=0.35, linestyle="--")
    ax.set_axisbelow(True)
    ax.spines[["top", "right"]].set_visible(False)

    plt.tight_layout(pad=2.5)
    plt.savefig(save_path, dpi=150, bbox_inches="tight",
                facecolor="white")
    plt.close()
    print(f"  Chart saved  → {save_path}")


# ================================================================
#  8. EXCEL EXPORT (3 sheets, colour-coded)
#     Sheet 1: Results  — full predictions with scam_score
#     Sheet 2: Metrics  — accuracy / precision / recall / F1
#     Sheet 3: Confusion Matrix
# ================================================================
def export_excel(
    results:   pd.DataFrame,
    metrics:   dict,
    cm:        np.ndarray,
    save_path: str,
) -> None:
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        results.to_excel(writer, sheet_name="Results", index=False)
        wb  = writer.book

        # ── Sheet 1: Results ────────────────────────────────
        ws = writer.sheets["Results"]

        # Styles
        hdr_fill = PatternFill("solid", fgColor="1F3864")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        thin_side = Side(style="thin", color="D5D5D5")
        cell_border = Border(left=thin_side,  right=thin_side,
                             top=thin_side,   bottom=thin_side)
        spam_fill  = PatternFill("solid", fgColor="FFE0E0")
        ham_fill   = PatternFill("solid", fgColor="E0F7E9")
        error_fill = PatternFill("solid", fgColor="FFF3CD")

        col_widths = {"A": 58, "B": 15, "C": 17,
                      "D": 19, "E": 17, "F": 13, "G": 8}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        # Header row
        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center")
            cell.border    = cell_border
        ws.row_dimensions[1].height = 22

        # Data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            actual    = row[1].value
            predicted = row[2].value
            wrong     = (actual != predicted)
            for cell in row:
                cell.border    = cell_border
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="left" if cell.column == 1 else "center",
                    wrap_text=(cell.column == 1),
                )
                if wrong:
                    cell.fill = error_fill
                elif actual == "spam":
                    cell.fill = spam_fill
                else:
                    cell.fill = ham_fill

        # Freeze top row
        ws.freeze_panes = "A2"

        # Colour scale on scam_score (col F)
        last_row = ws.max_row
        ws.conditional_formatting.add(
            f"F2:F{last_row}",
            ColorScaleRule(
                start_type  ="num", start_value =0,   start_color ="63BE7B",
                mid_type    ="num", mid_value   =50,  mid_color   ="FFEB84",
                end_type    ="num", end_value   =100, end_color   ="F8696B",
            )
        )

        # Bold misclassified predicted_label
        red_font  = Font(color="C0392B", bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=3, max_col=3):
            cell = row[0]
            actual_cell = ws.cell(row=cell.row, column=2)
            if cell.value != actual_cell.value:
                cell.font = red_font

        # ── Sheet 2: Metrics ────────────────────────────────
        metrics_df = pd.DataFrame({
            "Metric"   : ["Accuracy", "Precision", "Recall", "F1-Score"],
            "Score"    : [round(v, 6) for v in metrics.values()],
            "Score (%)" : [f"{v * 100:.2f}%" for v in metrics.values()],
            "Interpretation": [
                "Overall correctness",
                "Of predicted spam, how many were actually spam",
                "Of actual spam, how many were caught",
                "Harmonic mean of Precision & Recall",
            ],
        })
        metrics_df.to_excel(writer, sheet_name="Metrics", index=False)
        ws2 = writer.sheets["Metrics"]
        for col, w in zip(["A","B","C","D"], [14, 12, 12, 42]):
            ws2.column_dimensions[col].width = w
        for cell in ws2[1]:
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center")

        # ── Sheet 3: Confusion Matrix ────────────────────────
        cm_df = pd.DataFrame(
            [
                ["Actual Ham",  int(cm[0, 0]), int(cm[0, 1])],
                ["Actual Spam", int(cm[1, 0]), int(cm[1, 1])],
            ],
            columns=["", "Predicted Ham", "Predicted Spam"],
        )
        cm_df.to_excel(writer, sheet_name="Confusion Matrix", index=False)
        ws3 = writer.sheets["Confusion Matrix"]
        for col in ["A", "B", "C"]:
            ws3.column_dimensions[col].width = 18
        for cell in ws3[1]:
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center")

    print(f"  Excel saved  → {save_path}")


# ================================================================
#  9. SAVE / LOAD ARTIFACTS
# ================================================================
def save_artifacts(
    model:      XGBClassifier,
    vectorizer: TfidfVectorizer,
    model_path: str = MODEL_PATH,
    tfidf_path: str = TFIDF_PATH,
) -> None:
    with open(model_path, "wb") as f:
        pickle.dump(model, f)
    with open(tfidf_path, "wb") as f:
        pickle.dump(vectorizer, f)
    print(f"  Model saved  → {model_path}  "
          f"({os.path.getsize(model_path)/1024:.1f} KB)")
    print(f"  TF-IDF saved → {tfidf_path}  "
          f"({os.path.getsize(tfidf_path)/1024:.1f} KB)")


def load_artifacts(
    model_path: str = MODEL_PATH,
    tfidf_path: str = TFIDF_PATH,
) -> tuple[XGBClassifier, TfidfVectorizer]:
    with open(model_path, "rb") as f:
        model = pickle.load(f)
    with open(tfidf_path, "rb") as f:
        vectorizer = pickle.load(f)
    return model, vectorizer


# ================================================================
#  10. REUSABLE INFERENCE — analyze_message()
#
#  Usage:
#      result = analyze_message("Win Rs.1 lakh! Call now.")
#      # → {"prediction": "spam", "spam_probability": 0.99,
#      #     "ham_probability": 0.01, "scam_score": 99.0}
#
#  The first call auto-loads the saved model & vectoriser.
#  Subsequent calls reuse the in-memory objects (fast).
# ================================================================
_CACHE: dict = {}   # module-level cache for loaded artifacts

def analyze_message(
    message:    str,
    model_path: str = MODEL_PATH,
    tfidf_path: str = TFIDF_PATH,
) -> dict:
    """
    Classify a single raw SMS message.

    Parameters
    ----------
    message    : str — raw SMS text
    model_path : path to the saved XGBoost model (.pkl)
    tfidf_path : path to the saved TF-IDF vectoriser (.pkl)

    Returns
    -------
    dict
        {
            "prediction"       : "spam" | "ham",
            "spam_probability" : float  (0.0 – 1.0),
            "ham_probability"  : float  (0.0 – 1.0),
            "scam_score"       : float  (0   – 100 ),
        }
    """
    # Lazy-load once, reuse on every subsequent call
    if "model" not in _CACHE:
        _CACHE["model"], _CACHE["vectorizer"] = \
            load_artifacts(model_path, tfidf_path)

    vec    = _CACHE["vectorizer"].transform([str(message).strip()])
    probs  = _CACHE["model"].predict_proba(vec)[0]
    spam_p = float(round(probs[1], 6))
    ham_p  = float(round(probs[0], 6))

    return {
        "prediction"       : "spam" if spam_p >= 0.50 else "ham",
        "spam_probability" : spam_p,
        "ham_probability"  : ham_p,
        "scam_score"       : round(spam_p * 100, 2),
    }


def analyze_batch(
    messages:   list[str],
    model_path: str = MODEL_PATH,
    tfidf_path: str = TFIDF_PATH,
) -> pd.DataFrame:
    """
    Classify a list of messages in one vectorised pass.

    Returns a DataFrame with the same columns as analyze_message().
    """
    if "model" not in _CACHE:
        _CACHE["model"], _CACHE["vectorizer"] = \
            load_artifacts(model_path, tfidf_path)

    cleaned = [str(m).strip() for m in messages]
    vecs    = _CACHE["vectorizer"].transform(cleaned)
    probs   = _CACHE["model"].predict_proba(vecs)

    return pd.DataFrame({
        "message"          : cleaned,
        "prediction"       : ["spam" if p >= 0.5 else "ham" for p in probs[:, 1]],
        "spam_probability" : np.round(probs[:, 1], 6),
        "ham_probability"  : np.round(probs[:, 0], 6),
        "scam_score"       : np.round(probs[:, 1] * 100, 2),
    })


# ================================================================
#  11. MAIN PIPELINE ORCHESTRATOR
# ================================================================
def run_pipeline(
    csv_path:            str  = CSV_PATH,
    use_gridsearch:      bool = False,
    save_excel:          bool = True,
    save_plot:           bool = True,
    save_model_artifacts:bool = True,
) -> tuple[XGBClassifier, TfidfVectorizer, pd.DataFrame, dict]:
    """
    End-to-end pipeline: load → split → vectorise → train
                         → evaluate → scam_score → export.

    Parameters
    ----------
    csv_path             : path to CSV with 'label' and 'message' columns
    use_gridsearch       : run GridSearchCV hyperparameter tuning
    save_excel           : write spam_detection_results.xlsx
    save_plot            : write spam_model_report.png
    save_model_artifacts : pickle model + vectoriser

    Returns
    -------
    (model, vectorizer, results_df, metrics_dict)
    """
    t_start = time.time()
    print("=" * 62)
    print("   SMS SPAM DETECTION PIPELINE  —  XGBoost + TF-IDF")
    print("=" * 62)

    # ── Step 1: Load ─────────────────────────────────────────
    print("\n[1/9] Loading data…")
    df = load_data(csv_path)

    # ── Step 2: Split 80/20 ──────────────────────────────────
    print(f"\n[2/9] Stratified train/test split ({int((1-TEST_SIZE)*100)}/{int(TEST_SIZE*100)})…")
    X = df[TEXT_COL]
    y = df["label_enc"]

    (X_train, X_test,
     y_train, y_test,
     train_idx, test_idx) = train_test_split(
        X, y, np.arange(len(df)),
        test_size    = TEST_SIZE,
        random_state = RANDOM_STATE,
        stratify     = y,
    )
    print(f"  Train: {len(X_train):,}  |  Test: {len(X_test):,}")

    # ── Step 3: TF-IDF ───────────────────────────────────────
    print("\n[3/9] Fitting TF-IDF vectoriser (unigrams + bigrams)…")
    vectorizer, X_tr, X_te = build_tfidf(X_train, X_test)

    # ── Step 4: Train ────────────────────────────────────────
    print(f"\n[4/9] Training XGBoost (gridsearch={use_gridsearch})…")
    model = train_model(X_tr, y_train, use_gridsearch=use_gridsearch)

    # ── Step 5: Evaluate ─────────────────────────────────────
    print("\n[5/9] Evaluating on held-out test set…")
    y_pred, y_prob, metrics, cm = evaluate(model, X_te, y_test)

    # ── Step 6: scam_score for entire dataset ────────────────
    print("\n[6/9] Computing scam_score for all records…")
    results = build_results_df(df, vectorizer, model, train_idx, test_idx)

    # ── Step 7: Excel export ─────────────────────────────────
    if save_excel:
        print("\n[7/9] Writing Excel workbook…")
        export_excel(results, metrics, cm, OUTPUT_XLSX)

    # ── Step 8: Visualisation ────────────────────────────────
    if save_plot:
        print("\n[8/9] Generating evaluation chart…")
        plot_report(cm, metrics, model, vectorizer,
                    y_test.values, y_prob, PLOT_PATH)

    # ── Step 9: Save artifacts ───────────────────────────────
    if save_model_artifacts:
        print("\n[9/9] Saving model artifacts…")
        save_artifacts(model, vectorizer)

    elapsed = time.time() - t_start
    print(f"\n{'='*62}")
    print(f"   Pipeline complete in {elapsed:.1f}s")
    print(f"   Outputs: {OUTPUT_XLSX} | {PLOT_PATH}")
    print(f"   Artifacts: {MODEL_PATH} | {TFIDF_PATH}")
    print(f"{'='*62}\n")

    return model, vectorizer, results, metrics


# ================================================================
#  ENTRY POINT
# ================================================================
if __name__ == "__main__":

    # ── Run pipeline ─────────────────────────────────────────
    model, vectorizer, results, metrics = run_pipeline(
        csv_path             = CSV_PATH,
        use_gridsearch       = False,   # ← True = GridSearchCV tuning
        save_excel           = True,
        save_plot            = True,
        save_model_artifacts = True,
    )

    # ── Preview results DataFrame ────────────────────────────
    print("\n── Sample Results (first 5 rows) ──────────────────")
    pd.set_option("display.max_colwidth", 55)
    pd.set_option("display.width", 120)
    print(results[["message", "actual_label", "predicted_label",
                   "scam_score"]].head())

    # ── Demo: analyze_message() ──────────────────────────────
    print("\n── analyze_message() — live inference demo ─────────")
    demo = [
        ("WINNER ALERT! Your mobile number selected for Rs.1 Lakh gift. Visit http://amzn-reward.xyz now!", "spam"),
        ("HDFC Bank: Rs.2,500 debited from A/c XX4521 on 27-Nov via UPI. Avl bal: Rs.18,234. -HDFC", "ham"),
        ("Hey, are you free this evening? Let's catch up for chai.", "ham"),
        ("Your SBI account has been suspended. Update KYC at http://sbi-kyc.xyz immediately.", "spam"),
        ("Your OTP for IRCTC login is 447821. Please do not share. -IRCTC", "ham"),
        ("TRAI notice: Your SIM will be permanently blocked in 24 hours. Verify at http://trai-verify.xyz", "spam"),
        ("Reminder: Your dental checkup is on Nov 28 at 11 AM. Dr. Meena Dental Care, MG Road.", "ham"),
        ("Earn Rs.5,000/day liking YouTube videos! 2 hours work from home. Join: http://yt-earn.xyz", "spam"),
    ]
    print(f"  {'Result':<10} {'scam_score':>12}  {'Expected':<8}  Message")
    print("  " + "─" * 78)
    correct = 0
    for msg, expected in demo:
        r = analyze_message(msg)
        flag   = "🚨 SPAM" if r["prediction"] == "spam" else "✅ HAM "
        tick   = "✓" if r["prediction"] == expected else "✗"
        correct += (r["prediction"] == expected)
        print(f"  {flag}  {r['scam_score']:>8.1f}   [{tick}] {expected:<6}  {msg[:55]}…")

    print(f"\n  Demo accuracy: {correct}/{len(demo)} = {correct/len(demo)*100:.0f}%")

    # ── Demo: analyze_batch() ────────────────────────────────
    print("\n── analyze_batch() — vectorised batch inference ────")
    batch_msgs = [m for m, _ in demo]
    batch_df   = analyze_batch(batch_msgs)
    print(batch_df[["prediction", "scam_score", "message"]].to_string(index=False))
